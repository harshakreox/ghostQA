"""
Data Dictionary Parser
Parses Excel/CSV files containing field definitions for validation test generation
"""

import os
import csv
import io
from typing import List, Dict, Optional, Any
from pydantic import BaseModel


class FieldDefinition(BaseModel):
    """Definition of a single field from data dictionary"""
    field_name: str
    data_type: str = "string"  # string, number, email, date, boolean, etc.
    field_type: Optional[str] = None  # UI field type: text, dropdown, checkbox, etc.
    required: bool = False
    editable: bool = True
    min_length: Optional[int] = None
    max_length: Optional[int] = None
    min_value: Optional[float] = None
    max_value: Optional[float] = None
    allowed_values: Optional[List[str]] = None  # For dropdowns/enums
    pattern: Optional[str] = None  # Regex pattern or business rule
    description: Optional[str] = None
    page_name: Optional[str] = None  # Page or form this field belongs to
    section: Optional[str] = None  # Section within the page
    roles: Optional[str] = None  # User roles that can access this field


class DataDictionary(BaseModel):
    """Complete data dictionary with all field definitions"""
    name: str = "Data Dictionary"
    fields: List[FieldDefinition] = []
    # NEW: Raw data for AI-driven parsing
    headers: List[str] = []
    raw_rows: List[List[str]] = []

    def estimate_tokens(self) -> int:
        """Estimate token count for the raw table (rough estimate: ~4 chars per token)"""
        if self.headers and self.raw_rows:
            raw_table = self.to_raw_table()
            # Rough estimation: ~4 characters per token for English text
            return len(raw_table) // 4
        elif self.fields:
            context = self.to_prompt_context()
            return len(context) // 4
        return 0

    def get_batch(self, start_idx: int, batch_size: int) -> 'DataDictionary':
        """Get a batch of rows for chunked processing"""
        return DataDictionary(
            name=self.name,
            headers=self.headers,
            raw_rows=self.raw_rows[start_idx:start_idx + batch_size],
            fields=[]
        )

    def to_prompt_context(self) -> str:
        """Convert to context string for AI prompt - legacy method"""
        lines = [f"DATA DICTIONARY: {self.name}", "=" * 50, ""]

        for field in self.fields:
            lines.append(f"Field: {field.field_name}")
            lines.append(f"  - Data Type: {field.data_type}")
            lines.append(f"  - Required: {'Yes' if field.required else 'No'}")
            lines.append("")

        return "\n".join(lines)

    def to_raw_table(self) -> str:
        """Convert to compact table format for AI - optimized for minimal tokens"""
        if not self.headers or not self.raw_rows:
            return self.to_prompt_context()

        # Use compact format: pipe-separated, minimal whitespace
        lines = [
            f"[{self.name}] {len(self.raw_rows)} entries",
            "COLS: " + "|".join(h.strip() for h in self.headers),
            "DATA:"
        ]

        # Compact row format - no "Row N:" prefix, just data
        for row in self.raw_rows:
            # Pad row and join with pipes, trim empty trailing cells
            padded_row = row + [''] * (len(self.headers) - len(row))
            # Remove trailing empty cells to save tokens
            while padded_row and not padded_row[-1]:
                padded_row.pop()
            if padded_row:
                lines.append("|".join(str(cell).strip() if cell else '' for cell in padded_row))

        return "\n".join(lines)


# Column name mappings (case-insensitive) - expanded for flexibility
COLUMN_MAPPINGS = {
    'field_name': [
        'field_name', 'field', 'name', 'column', 'column_name', 'attribute', 'property',
        'fieldname', 'field_id', 'element', 'element_name', 'input', 'input_name',
        'label', 'field_label', 'control', 'control_name', 'parameter', 'param',
        'variable', 'var', 'key', 'identifier', 'attr', 'attribute_name',
        # Common variations
        'field_name/label', 'fieldname/label', 'field_name_label', 'data_element'
    ],
    'data_type': [
        'data_type', 'type', 'datatype', 'field_type', 'format', 'data_format',
        'value_type', 'input_type', 'control_type', 'dtype', 'kind'
    ],
    'field_type': [
        'field_type', 'input_type', 'control_type', 'ui_type', 'widget',
        'control', 'element_type'
    ],
    'required': [
        'required', 'mandatory', 'is_required', 'nullable', 'not_null', 'optional',
        'is_mandatory', 'is_optional', 'null', 'allow_null', 'required_field',
        'compulsory', 'must_have', 'needed',
        # Y/N variations
        'mandatory_(y/n)', 'mandatory_y/n', 'required_(y/n)', 'required_y/n'
    ],
    'min_length': [
        'min_length', 'minlength', 'min_len', 'minimum_length', 'min_chars',
        'minimum_chars', 'min_size', 'length_min'
    ],
    'max_length': [
        'max_length', 'maxlength', 'max_len', 'maximum_length', 'max_chars',
        'maximum_chars', 'max_size', 'length_max', 'length', 'size', 'char_limit'
    ],
    'min_value': [
        'min_value', 'minvalue', 'min', 'minimum', 'min_val', 'range_min',
        'lower_bound', 'lower_limit', 'from_value'
    ],
    'max_value': [
        'max_value', 'maxvalue', 'max', 'maximum', 'max_val', 'range_max',
        'upper_bound', 'upper_limit', 'to_value'
    ],
    'allowed_values': [
        'allowed_values', 'values', 'options', 'enum', 'choices', 'valid_values',
        'possible_values', 'accepted_values', 'dropdown', 'dropdown_values',
        'list_values', 'selection', 'pick_list', 'lookup', 'domain',
        'default_value', 'default'
    ],
    'pattern': [
        'pattern', 'regex', 'format_pattern', 'validation_pattern', 'regexp',
        'regular_expression', 'format_regex', 'input_pattern', 'mask',
        # Business rule variations
        'business_rule', 'business_rule/validation', 'validation', 'validation_rule',
        'rule', 'business_logic'
    ],
    'description': [
        'description', 'desc', 'comment', 'notes', 'remarks', 'definition',
        'explanation', 'details', 'info', 'information', 'help_text', 'tooltip',
        'business_rule', 'business_rule/validation', 'section'
    ],
    # Extra columns to capture
    'page_name': [
        'page', 'page_name', 'form', 'form_name', 'page/form_name', 'page/form',
        'screen', 'screen_name', 'module'
    ],
    'section': [
        'section', 'section_name', 'group', 'category', 'area', 'panel'
    ],
    'editable': [
        'editable', 'editable_(y/n)', 'editable_y/n', 'readonly', 'read_only',
        'is_editable', 'can_edit', 'modifiable'
    ],
    'roles': [
        'role', 'roles', 'role(s)', 'user_role', 'access', 'permissions'
    ]
}


def find_column_mapping(headers: List[str], verbose: bool = False) -> Dict[str, Optional[int]]:
    """Map standard field names to column indices with flexible matching"""
    mapping = {}

    # Normalize headers: lowercase, remove spaces/underscores/hyphens
    headers_normalized = []
    for h in headers:
        if h:
            normalized = h.lower().strip().replace(' ', '_').replace('-', '_')
            headers_normalized.append(normalized)
        else:
            headers_normalized.append('')

    if verbose:
        print(f"   Headers found: {headers}")
        print(f"   Normalized: {headers_normalized}")

    for field, aliases in COLUMN_MAPPINGS.items():
        mapping[field] = None

        # Try exact match first
        for alias in aliases:
            if alias in headers_normalized:
                mapping[field] = headers_normalized.index(alias)
                if verbose:
                    print(f"   Matched '{field}' to column {mapping[field]} ('{headers[mapping[field]]}')")
                break

        # If no exact match, try partial/contains matching
        if mapping[field] is None:
            for idx, header in enumerate(headers_normalized):
                if header:
                    for alias in aliases:
                        # Check if alias is contained in header or vice versa
                        if alias in header or header in alias:
                            mapping[field] = idx
                            if verbose:
                                print(f"   Partial matched '{field}' to column {idx} ('{headers[idx]}')")
                            break
                    if mapping[field] is not None:
                        break

    # Special fallback: if no field_name found, use first non-empty text column
    if mapping['field_name'] is None and len(headers) > 0:
        # Use first column as field_name
        mapping['field_name'] = 0
        if verbose:
            print(f"   Fallback: using first column as field_name ('{headers[0]}')")

    return mapping


def parse_boolean(value: Any) -> bool:
    """Parse various boolean representations"""
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.lower().strip() in ['yes', 'true', '1', 'y', 'required', 'mandatory']
    if isinstance(value, (int, float)):
        return bool(value)
    return False


def parse_int(value: Any) -> Optional[int]:
    """Parse integer value"""
    if value is None or value == '':
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def parse_float(value: Any) -> Optional[float]:
    """Parse float value"""
    if value is None or value == '':
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def parse_list(value: Any) -> Optional[List[str]]:
    """Parse comma-separated list"""
    if value is None or value == '':
        return None
    if isinstance(value, list):
        return value
    if isinstance(value, str):
        items = [item.strip() for item in value.split(',') if item.strip()]
        return items if items else None
    return None


def parse_csv_content(content: str, filename: str = "data_dictionary") -> DataDictionary:
    """Parse CSV content into DataDictionary"""

    # Try to detect delimiter (comma, semicolon, tab)
    first_line = content.split('\n')[0] if content else ''
    delimiter = ','
    if ';' in first_line and ',' not in first_line:
        delimiter = ';'
    elif '\t' in first_line:
        delimiter = '\t'

    reader = csv.reader(io.StringIO(content), delimiter=delimiter)
    rows = list(reader)

    if len(rows) < 2:
        raise ValueError("CSV must have at least a header row and one data row")

    headers = rows[0]
    print(f"   CSV Headers: {headers}")

    mapping = find_column_mapping(headers, verbose=True)

    print(f"   Column mapping: {mapping}")

    fields = []
    for row in rows[1:]:
        if not row or not any(cell.strip() for cell in row if cell):
            continue  # Skip empty rows

        def get_value(key: str) -> Any:
            idx = mapping.get(key)
            if idx is not None and idx < len(row):
                val = row[idx]
                return val.strip() if val else None
            return None

        field_name = get_value('field_name')
        if not field_name:
            continue

        # Parse editable field (default to True, but check for N/No/False)
        editable_val = get_value('editable')
        editable = True
        if editable_val:
            editable = editable_val.lower().strip() not in ['n', 'no', 'false', '0', 'readonly', 'read-only']

        field = FieldDefinition(
            field_name=field_name,
            data_type=get_value('data_type') or 'string',
            field_type=get_value('field_type'),
            required=parse_boolean(get_value('required')),
            editable=editable,
            min_length=parse_int(get_value('min_length')),
            max_length=parse_int(get_value('max_length')),
            min_value=parse_float(get_value('min_value')),
            max_value=parse_float(get_value('max_value')),
            allowed_values=parse_list(get_value('allowed_values')),
            pattern=get_value('pattern'),
            description=get_value('description'),
            page_name=get_value('page_name'),
            section=get_value('section'),
            roles=get_value('roles')
        )
        fields.append(field)

    if not fields:
        raise ValueError("No valid field definitions found in CSV. Make sure your file has a header row with recognizable column names.")

    # Extract name from filename
    name = os.path.splitext(os.path.basename(filename))[0].replace('_', ' ').title()

    print(f"   Successfully parsed {len(fields)} fields")

    return DataDictionary(name=name, fields=fields)


def parse_excel_content(file_path: str) -> DataDictionary:
    """Parse Excel file into DataDictionary"""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel parsing. Run: pip install openpyxl")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))

    if len(rows) < 2:
        raise ValueError("Excel must have at least a header row and one data row")

    headers = [str(h) if h else '' for h in rows[0]]
    print(f"   Excel Headers: {headers}")

    mapping = find_column_mapping(headers, verbose=True)

    print(f"   Column mapping: {mapping}")

    fields = []
    for row in rows[1:]:
        if not row or not any(cell for cell in row if cell):
            continue  # Skip empty rows

        def get_value(key: str) -> Any:
            idx = mapping.get(key)
            if idx is not None and idx < len(row):
                val = row[idx]
                return str(val).strip() if val is not None else None
            return None

        field_name = get_value('field_name')
        if not field_name:
            continue

        # Parse editable field (default to True, but check for N/No/False)
        editable_val = get_value('editable')
        editable = True
        if editable_val:
            editable = editable_val.lower().strip() not in ['n', 'no', 'false', '0', 'readonly', 'read-only']

        field = FieldDefinition(
            field_name=field_name,
            data_type=get_value('data_type') or 'string',
            field_type=get_value('field_type'),
            required=parse_boolean(get_value('required')),
            editable=editable,
            min_length=parse_int(get_value('min_length')),
            max_length=parse_int(get_value('max_length')),
            min_value=parse_float(get_value('min_value')),
            max_value=parse_float(get_value('max_value')),
            allowed_values=parse_list(get_value('allowed_values')),
            pattern=get_value('pattern'),
            description=get_value('description'),
            page_name=get_value('page_name'),
            section=get_value('section'),
            roles=get_value('roles')
        )
        fields.append(field)

    if not fields:
        raise ValueError("No valid field definitions found in Excel. Make sure your file has a header row with recognizable column names.")

    # Extract name from filename
    name = os.path.splitext(os.path.basename(file_path))[0].replace('_', ' ').title()

    print(f"   Successfully parsed {len(fields)} fields")

    return DataDictionary(name=name, fields=fields)


def parse_raw_csv(file_path: str) -> DataDictionary:
    """Parse CSV file and return raw headers + rows for AI to interpret"""
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # Try to detect delimiter
    first_line = content.split('\n')[0] if content else ''
    delimiter = ','
    if ';' in first_line and ',' not in first_line:
        delimiter = ';'
    elif '\t' in first_line:
        delimiter = '\t'

    reader = csv.reader(io.StringIO(content), delimiter=delimiter)
    rows = list(reader)

    if len(rows) < 2:
        raise ValueError("File must have at least a header row and one data row")

    headers = [h.strip() if h else '' for h in rows[0]]
    data_rows = []

    for row in rows[1:]:
        if row and any(cell.strip() for cell in row if cell):
            cleaned_row = [cell.strip() if cell else '' for cell in row]
            data_rows.append(cleaned_row)

    name = os.path.splitext(os.path.basename(file_path))[0].replace('_', ' ').title()

    print(f"   Raw parse: {len(headers)} columns, {len(data_rows)} rows")

    return DataDictionary(
        name=name,
        headers=headers,
        raw_rows=data_rows
    )


def parse_raw_excel(file_path: str) -> DataDictionary:
    """Parse Excel file and return raw headers + rows for AI to interpret"""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel parsing. Run: pip install openpyxl")

    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))

    if len(rows) < 2:
        raise ValueError("File must have at least a header row and one data row")

    headers = [str(h).strip() if h else '' for h in rows[0]]
    data_rows = []

    for row in rows[1:]:
        if row and any(cell for cell in row if cell):
            cleaned_row = [str(cell).strip() if cell else '' for cell in row]
            data_rows.append(cleaned_row)

    name = os.path.splitext(os.path.basename(file_path))[0].replace('_', ' ').title()

    print(f"   Raw parse: {len(headers)} columns, {len(data_rows)} rows")

    return DataDictionary(
        name=name,
        headers=headers,
        raw_rows=data_rows
    )


def parse_data_dictionary_raw(file_path: str) -> DataDictionary:
    """Parse data dictionary as raw data - let AI interpret the structure"""
    ext = os.path.splitext(file_path)[1].lower()

    if ext == '.csv':
        return parse_raw_csv(file_path)
    elif ext in ['.xlsx', '.xls']:
        return parse_raw_excel(file_path)
    else:
        raise ValueError(f"Unsupported file format: {ext}. Use .csv or .xlsx")


def parse_data_dictionary_file(file_path: str) -> DataDictionary:
    """Parse data dictionary from file (auto-detect format)"""
    ext = os.path.splitext(file_path)[1].lower()

    if ext == '.csv':
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return parse_csv_content(content, file_path)

    elif ext in ['.xlsx', '.xls']:
        return parse_excel_content(file_path)

    else:
        raise ValueError(f"Unsupported file format: {ext}. Use .csv or .xlsx")
